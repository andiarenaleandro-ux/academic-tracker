from datetime import date, datetime
from io import BytesIO
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.models.carrera import Carrera
from app.models.semestre import Semestre
from app.models.materia import Materia
from app.models.evaluacion import Evaluacion
from app.models.clase import Clase
from app.models.config_asistencia import ConfigAsistencia
from app.models.correlatividad import Correlatividad


@dataclass
class RowError:
    sheet: str
    row: int
    message: str


@dataclass
class SheetResult:
    created: int = 0
    updated: int = 0
    errors: list[RowError] = field(default_factory=list)


@dataclass
class ImportResult:
    carreras: SheetResult = field(default_factory=SheetResult)
    semestres: SheetResult = field(default_factory=SheetResult)
    materias: SheetResult = field(default_factory=SheetResult)
    evaluaciones: SheetResult = field(default_factory=SheetResult)
    clases: SheetResult = field(default_factory=SheetResult)
    asistencia: SheetResult = field(default_factory=SheetResult)

    def total_created(self) -> int:
        return sum(
            s.created for s in [self.carreras, self.semestres, self.materias, self.evaluaciones, self.clases, self.asistencia]
        )

    def total_errors(self) -> int:
        return sum(
            len(s.errors) for s in [self.carreras, self.semestres, self.materias, self.evaluaciones, self.clases, self.asistencia]
        )


SHEET_COLUMNS = {
    "Carreras": {"nombre", "escala_nota_min", "escala_nota_max", "nota_aprobacion"},
    "Semestres": {"numero", "anio", "fecha_inicio", "fecha_fin", "carrera"},
    "Materias": {"nombre", "semestre", "creditos", "profesor", "estado", "duracion"},
    "Evaluaciones": {"materia", "tipo", "fecha", "peso", "nota_obtenida", "nota_simulada", "notas"},
    "Clases": {"materia", "fecha", "asistio", "tema", "notas"},
    "Asistencia": {"materia", "asistencia_minima_pct"},
}


def _cell_value(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "value"):
        v = v.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_bool(val: str) -> bool:
    return val.lower() in ("true", "verdadero", "si", "1", "yes")


def _parse_float(val: str) -> float | None:
    try:
        return float(val.replace(",", "."))
    except (ValueError, AttributeError):
        return None


def _header_map(headers: list[str]) -> dict[str, int]:
    return {h.strip().lower(): i for i, h in enumerate(headers) if h}


def _get_row_dict(headers: dict[str, int], row: tuple) -> dict[str, str]:
    return {col: _cell_value(row[idx]) if idx < len(row) else "" for col, idx in headers.items()}


def _resolve_carrera(db: Session, nombre: str) -> Carrera | None:
    return db.query(Carrera).filter(Carrera.nombre == nombre).first()


def _resolve_semestre(db: Session, carrera_id: int, numero: int, anio: int) -> Semestre | None:
    return db.query(Semestre).filter(
        Semestre.carrera_id == carrera_id,
        Semestre.numero == numero,
        Semestre.anio == anio,
    ).first()


def _resolve_materia(db: Session, semestre_id: int, nombre: str) -> Materia | None:
    return db.query(Materia).filter(
        Materia.semestre_id == semestre_id,
        Materia.nombre == nombre,
    ).first()


def _parse_semestre_key(val: str) -> tuple[int, int] | None:
    parts = val.split("-")
    if len(parts) == 2:
        try:
            return int(parts[0]), int(parts[1])
        except ValueError:
            return None
    return None


def import_carreras(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"nombre"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Carreras", row=0, message="Faltan columnas requeridas: nombre"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        nombre = d.get("nombre", "")
        if not nombre:
            result.errors.append(RowError(sheet="Carreras", row=i, message="nombre vacío"))
            continue
        existing = _resolve_carrera(db, nombre)
        if existing:
            if d.get("escala_nota_min"):
                existing.escala_nota_min = int(d["escala_nota_min"])
            if d.get("escala_nota_max"):
                existing.escala_nota_max = int(d["escala_nota_max"])
            if d.get("nota_aprobacion"):
                existing.nota_aprobacion = float(d["nota_aprobacion"])
            result.updated += 1
        else:
            carrera = Carrera(
                nombre=nombre,
                escala_nota_min=int(d["escala_nota_min"]) if d.get("escala_nota_min") else 0,
                escala_nota_max=int(d["escala_nota_max"]) if d.get("escala_nota_max") else 10,
                nota_aprobacion=float(d["nota_aprobacion"]) if d.get("nota_aprobacion") else 4.0,
            )
            db.add(carrera)
            result.created += 1
    db.commit()


def import_semestres(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"numero", "anio", "fecha_inicio", "fecha_fin", "carrera"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Semestres", row=0, message="Faltan columnas requeridas"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        carrera_nombre = d.get("carrera", "")
        carrera = _resolve_carrera(db, carrera_nombre)
        if not carrera:
            result.errors.append(RowError(sheet="Semestres", row=i, message=f"Carrera '{carrera_nombre}' no encontrada"))
            continue
        try:
            numero = int(d["numero"])
            anio = int(d["anio"])
        except (ValueError, KeyError):
            result.errors.append(RowError(sheet="Semestres", row=i, message="numero o anio inválidos"))
            continue
        existing = _resolve_semestre(db, carrera.id, numero, anio)
        if existing:
            result.updated += 1
        else:
            try:
                fecha_inicio = date.fromisoformat(d["fecha_inicio"])
                fecha_fin = date.fromisoformat(d["fecha_fin"])
            except (ValueError, KeyError):
                result.errors.append(RowError(sheet="Semestres", row=i, message="fechas inválidas"))
                continue
            semestre = Semestre(
                carrera_id=carrera.id,
                numero=numero,
                anio=anio,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
            )
            db.add(semestre)
            result.created += 1
    db.commit()


def import_materias(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"nombre", "semestre"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Materias", row=0, message="Faltan columnas requeridas: nombre, semestre"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        nombre = d.get("nombre", "")
        semestre_key = d.get("semestre", "")
        if not nombre or not semestre_key:
            result.errors.append(RowError(sheet="Materias", row=i, message="nombre o semestre vacío"))
            continue
        parsed = _parse_semestre_key(semestre_key)
        if not parsed:
            result.errors.append(RowError(sheet="Materias", row=i, message=f"Formato de semestre inválido: '{semestre_key}'. Esperado: 'num-anio'"))
            continue
        num, anio = parsed
        semestre = db.query(Semestre).filter(
            Semestre.numero == num, Semestre.anio == anio,
        ).first()
        if not semestre:
            result.errors.append(RowError(sheet="Materias", row=i, message=f"Semestre {semestre_key} no encontrado"))
            continue
        existing = _resolve_materia(db, semestre.id, nombre)
        if existing:
            if d.get("creditos"):
                existing.creditos = int(d["creditos"])
            if d.get("profesor"):
                existing.profesor = d["profesor"]
            if d.get("estado"):
                existing.estado = d["estado"]
            if d.get("duracion"):
                existing.duracion = d["duracion"]
            result.updated += 1
        else:
            materia = Materia(
                semestre_id=semestre.id,
                nombre=nombre,
                creditos=int(d["creditos"]) if d.get("creditos") else None,
                profesor=d.get("profesor") or None,
                estado=d.get("estado") or "cursando",
                duracion=d.get("duracion") or None,
            )
            db.add(materia)
            result.created += 1
    db.commit()


def import_evaluaciones(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"materia", "tipo", "fecha"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Evaluaciones", row=0, message="Faltan columnas requeridas"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        materia_nombre = d.get("materia", "")
        if not materia_nombre:
            result.errors.append(RowError(sheet="Evaluaciones", row=i, message="materia vacío"))
            continue
        materia = db.query(Materia).filter(Materia.nombre == materia_nombre).first()
        if not materia:
            result.errors.append(RowError(sheet="Evaluaciones", row=i, message=f"Materia '{materia_nombre}' no encontrada"))
            continue
        try:
            fecha = date.fromisoformat(d.get("fecha", ""))
        except (ValueError, TypeError):
            result.errors.append(RowError(sheet="Evaluaciones", row=i, message="fecha inválida"))
            continue
        evaluacion = Evaluacion(
            materia_id=materia.id,
            tipo=d.get("tipo", "parcial"),
            fecha=fecha,
            peso=_parse_float(d.get("peso")) or 1.0,
            nota_obtenida=_parse_float(d.get("nota_obtenida")),
            nota_simulada=_parse_float(d.get("nota_simulada")),
            notas=d.get("notas") or None,
        )
        db.add(evaluacion)
        result.created += 1
    db.commit()


def import_clases(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"materia", "fecha"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Clases", row=0, message="Faltan columnas requeridas"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        materia_nombre = d.get("materia", "")
        if not materia_nombre:
            result.errors.append(RowError(sheet="Clases", row=i, message="materia vacío"))
            continue
        materia = db.query(Materia).filter(Materia.nombre == materia_nombre).first()
        if not materia:
            result.errors.append(RowError(sheet="Clases", row=i, message=f"Materia '{materia_nombre}' no encontrada"))
            continue
        try:
            fecha = date.fromisoformat(d.get("fecha", ""))
        except (ValueError, TypeError):
            result.errors.append(RowError(sheet="Clases", row=i, message="fecha inválida"))
            continue
        clase = Clase(
            materia_id=materia.id,
            fecha=fecha,
            asistio=_parse_bool(d.get("asistio", "false")),
            tema=d.get("tema") or None,
            notas=d.get("notas") or None,
        )
        db.add(clase)
        result.created += 1
    db.commit()


def import_asistencia(sheet, db: Session, result: SheetResult):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    headers = _header_map([str(c or "") for c in rows[0]])
    required = {"materia", "asistencia_minima_pct"}
    if not required.issubset(headers.keys()):
        result.errors.append(RowError(sheet="Asistencia", row=0, message="Faltan columnas requeridas"))
        return

    for i, row in enumerate(rows[1:], start=2):
        d = _get_row_dict(headers, row)
        materia_nombre = d.get("materia", "")
        if not materia_nombre:
            result.errors.append(RowError(sheet="Asistencia", row=i, message="materia vacío"))
            continue
        materia = db.query(Materia).filter(Materia.nombre == materia_nombre).first()
        if not materia:
            result.errors.append(RowError(sheet="Asistencia", row=i, message=f"Materia '{materia_nombre}' no encontrada"))
            continue
        pct = _parse_float(d.get("asistencia_minima_pct"))
        if pct is None:
            result.errors.append(RowError(sheet="Asistencia", row=i, message="asistencia_minima_pct inválido"))
            continue
        existing = db.query(ConfigAsistencia).filter(ConfigAsistencia.materia_id == materia.id).first()
        if existing:
            existing.asistencia_minima_pct = pct
            result.updated += 1
        else:
            db.add(ConfigAsistencia(materia_id=materia.id, asistencia_minima_pct=pct))
            result.created += 1
    db.commit()


def import_plan_json(data: dict[str, Any], db: Session) -> dict:
    nombre_carrera = data.get("carrera", "Mi Carrera")
    carrera = _resolve_carrera(db, nombre_carrera)
    if not carrera:
        carrera = Carrera(nombre=nombre_carrera, escala_nota_min=0, escala_nota_max=10, nota_aprobacion=4)
        db.add(carrera)
        db.flush()

    materias_creadas: dict[str, int] = {}

    for sem_data in data.get("semestres", []):
        numero = int(sem_data["numero"])
        anio = 2026
        semestre = _resolve_semestre(db, carrera.id, numero, anio)
        if not semestre:
            fecha_inicio = date(anio, 1, 1)
            fecha_fin = date(anio, 12, 31)
            if numero % 2 == 1:
                fecha_inicio = date(anio, 3, 1)
                fecha_fin = date(anio, 7, 31)
            else:
                fecha_inicio = date(anio, 8, 1)
                fecha_fin = date(anio, 12, 31)
            semestre = Semestre(
                carrera_id=carrera.id, numero=numero, anio=anio,
                fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
            )
            db.add(semestre)
            db.flush()

        for nombre in sem_data.get("materias", []):
            nombre = nombre.strip()
            if not nombre:
                continue
            existing = _resolve_materia(db, semestre.id, nombre)
            if not existing:
                materia = Materia(semestre_id=semestre.id, nombre=nombre, estado="pendiente")
                db.add(materia)
                db.flush()
                materias_creadas[nombre] = materia.id
            else:
                materias_creadas[nombre] = existing.id

    for rel in data.get("correlativas", []):
        materia_nombre = rel.get("materia", "").strip()
        materia_id = materias_creadas.get(materia_nombre)
        if not materia_id:
            continue
        for req_nombre in rel.get("requisitos", []):
            req_nombre = req_nombre.strip()
            req_id = materias_creadas.get(req_nombre)
            if not req_id:
                continue
            existing = db.query(Correlatividad).filter(
                Correlatividad.materia_id == materia_id,
                Correlatividad.correlativa_id == req_id,
            ).first()
            if not existing:
                db.add(Correlatividad(materia_id=materia_id, correlativa_id=req_id))

    db.commit()
    return {
        "ok": True,
        "creados": len(materias_creadas),
        "errores": 0,
        "detalle": {
            "carrera": nombre_carrera,
            "materias_creadas": len(materias_creadas),
            "correlativas_creadas": len(data.get("correlativas", [])),
        },
    }


SHEET_IMPORTERS = {
    "Carreras": import_carreras,
    "Semestres": import_semestres,
    "Materias": import_materias,
    "Evaluaciones": import_evaluaciones,
    "Clases": import_clases,
    "Asistencia": import_asistencia,
}


def process_import(file_bytes: bytes, db: Session) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    result = ImportResult()

    sheet_names = {name.strip() for name in wb.sheetnames}
    available = SHEET_IMPORTERS.keys() & sheet_names

    for name in ["Carreras", "Semestres", "Materias", "Evaluaciones", "Clases", "Asistencia"]:
        if name not in sheet_names:
            continue
        ws = wb[name]
        importer = SHEET_IMPORTERS[name]
        sheet_result = getattr(result, name.lower())
        importer(ws, db, sheet_result)

    wb.close()
    return {
        "ok": result.total_errors() == 0,
        "creados": result.total_created(),
        "errores": result.total_errors(),
        "detalle": {
            "carreras": {"creados": result.carreras.created, "actualizados": result.carreras.updated, "errores": [str(e.message) for e in result.carreras.errors]},
            "semestres": {"creados": result.semestres.created, "actualizados": result.semestres.updated, "errores": [str(e.message) for e in result.semestres.errors]},
            "materias": {"creados": result.materias.created, "actualizados": result.materias.updated, "errores": [str(e.message) for e in result.materias.errors]},
            "evaluaciones": {"creados": result.evaluaciones.created, "actualizados": result.evaluaciones.updated, "errores": [str(e.message) for e in result.evaluaciones.errors]},
            "clases": {"creados": result.clases.created, "actualizados": result.clases.updated, "errores": [str(e.message) for e in result.clases.errors]},
            "asistencia": {"creados": result.asistencia.created, "actualizados": result.asistencia.updated, "errores": [str(e.message) for e in result.asistencia.errors]},
        },
    }
