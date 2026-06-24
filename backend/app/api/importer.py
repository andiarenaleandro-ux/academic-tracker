import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.db import get_db
from app.services.importer import process_import, import_plan_json

router = APIRouter()


class CorrelativaRel(BaseModel):
    materia: str
    requisitos: list[str]


class SemestrePlan(BaseModel):
    numero: int
    materias: list[str]


class PlanImport(BaseModel):
    carrera: str = "Mi Carrera"
    semestres: list[SemestrePlan]
    correlativas: list[CorrelativaRel] = []


@router.post("/import", status_code=200)
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo archivos .xlsx")
    contents = await file.read()
    if len(contents) == 0:
        raise HTTPException(400, "Archivo vacío")
    return process_import(contents, db)


@router.post("/import/plan", status_code=200)
def import_plan(data: PlanImport, db: Session = Depends(get_db)):
    return import_plan_json(data.model_dump(), db)


@router.get("/import/template")
def download_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    h_fill = PatternFill(start_color="3F3F5F", end_color="3F3F5F", fill_type="solid")
    h_font = Font(bold=True, color="C4B5FD")
    ex_font = Font(color="A1A1AA", italic=True)
    center = Alignment(horizontal="center")

    sheets = [
        ("Carreras", ["nombre", "escala_nota_min", "escala_nota_max", "nota_aprobacion"], [
            ["Mi Carrera", 0, 10, 4.0],
        ]),
        ("Semestres", ["numero", "anio", "fecha_inicio", "fecha_fin", "carrera"], [
            [1, 2025, "2025-03-01", "2025-07-31", "Mi Carrera"],
            [2, 2025, "2025-08-01", "2025-12-15", "Mi Carrera"],
        ]),
        # semestre = "numero-anio" (ej: 1-2025)
        # estado: cursando | aprobada | recursando | libre | pendiente
        # duracion: anual | cuatrimestral | semestral | bimestral | trimestral
        ("Materias", ["nombre", "semestre", "creditos", "profesor", "estado", "duracion"], [
            ["Matemática I", "1-2025", 6, "Prof. García", "cursando", "cuatrimestral"],
            ["Programación I", "1-2025", 4, "Prof. López", "cursando", "cuatrimestral"],
            ["Matemática II", "2-2025", 6, "Prof. García", "pendiente", "anual"],
        ]),
        # tipo: parcial | recuperatorio | tp | final | coloquio
        ("Evaluaciones", ["materia", "tipo", "fecha", "peso", "nota_obtenida", "nota_simulada", "notas"], [
            ["Matemática I", "parcial", "2025-05-10", 1, 7.5, "", "Primer parcial"],
        ]),
        # asistio: true | false
        ("Clases", ["materia", "fecha", "asistio", "tema", "notas"], [
            ["Matemática I", "2025-03-05", "true", "Números reales", ""],
        ]),
        ("Asistencia", ["materia", "asistencia_minima_pct"], [
            ["Matemática I", 75],
        ]),
    ]

    for name, headers, examples in sheets:
        ws = wb.create_sheet(name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 16)
        for row_i, example in enumerate(examples, 2):
            for col, val in enumerate(example, 1):
                ws.cell(row=row_i, column=col, value=val).font = ex_font
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_academic_tracker.xlsx"},
    )
